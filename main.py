# 地图绘制.py
# Use Tab to ident (BAD HABIT BUT ...)
# 使用Tab来缩进（坏习惯。。。）

import time
import sys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from tkinter import *
import matplotlib.pyplot as plt

class TrackEditor:
    def __init__(self):
        # 按键锁
        self.in_loop = 0 #开始是0 判断，防止连点两个功能卡死
        self.data1 = []
        self.data2 = []
        self.data3 = []
        #####
        self.Load_Data()
        self.root = Tk()
        self.root.title('赛道编辑器')
        self.root.geometry("+900+100")
        self.label = Label(self.root)
        self.label['text'] = '\n赛道编辑器\n\n'
        self.label.pack()
        #####
        self.Check = Button(self.root, text='查看锥桶位置', command=self.Matlab_Check)
        self.Txt = Text(self.root, height=3, width=20)
        self.Check.pack()
        self.Txt.pack()
        #####
        self.Add_blue = Button(self.root, text='  鼠标添加blue  ', command=self.Matlab_Add_Blue)
        self.Add_blue.pack()
        #####
        self.Add_red = Button(self.root, text='  鼠标添加red   ', command=self.Matlab_Add_Red)
        self.Add_red.pack()
        #####
        self.Input_x = Entry(self.root)
        self.Input_y = Entry(self.root)
        self.Input_type = Entry(self.root)
        self.Add = Button(self.root, text='  坐标添加  ', command=self.Matlab_Add)
        self.Input_x.pack()
        self.Input_y.pack()
        self.Input_type.pack()
        self.Add.pack()
        #####
        self.Delete = Button(self.root, text='  删除锥桶  ', command=self.Matlab_Delete)
        self.Delete.pack()
        #####
        self.Refresh = Button(self.root, text='    刷新    ', command=self.Refresh)
        self.Refresh.pack()
        #####
        self.Out = Button(self.root, text='    导出    ', command=self.Out_Data)
        self.Out.pack()
        #####
        self.b = Button(self.root, text='    退出    ', command=self.All_Close)
        self.b.pack()
        #####
        self.Help = Label(self.root)
        self.Help['text'] = '1.输入x坐标, y坐标, red或blue 来添加自定义锥筒\n'
        self.Help['text'] += '2.在使用其他功能前请先刷新，以免报错\n'
        self.Help['text'] += '3.当出现卡死的情况，果断关闭plot，一切就又好起来了（确信）\n'
        self.Help['text'] += 'Designed by FGH'
        self.Help.pack()
        #####

    # 导入
    def Load_Data(self):
        workbook = load_workbook('Map.xlsx')  # 找到需要xlsx文件的位置
        booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
        # 如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
        # sheets = workbook.get_sheet_names()  # 从名称获取sheet
        # booksheet = workbook.get_sheet_by_name(sheets[0])
        # 获取sheet页的行数据
        rows = booksheet.rows
        # 获取sheet页的列数据
        columns = booksheet.columns
        i = 0
        # 迭代所有的行
        for row in rows:
            i = i + 1
            line = [col.value for col in row]
            self.data1.append(booksheet.cell(row=i, column=1).value)  # 获取第i行1列的数据
            self.data2.append(booksheet.cell(row=i, column=2).value)  # 获取第i行2列的数据
            self.data3.append(booksheet.cell(row=i, column=3).value)  # 获取第i行3列的数据

    # 绘图函数
    def Matlab_Drawing(self):
        for i in range(len(self.data3)):
            if self.data3[i] == 1:
                plt.scatter(self.data1[i], self.data2[i],
                            color='blue', marker='.')
            else:
                plt.scatter(self.data1[i], self.data2[i],
                            color='red', marker='.')

    # 绘图函数的删除功能
    def Matlab_Delete(self):
        self.Refresh_Loop()
        while (1):
            self.Matlab_Drawing()
            try:
                [(m, n)] = plt.ginput(1)
            except ValueError:
                time.sleep(0.01)
                [(m, n)] = plt.ginput(1)
            for i in range(len(self.data3)):
                if abs(self.data1[i] - m) <= 0.5 and abs(self.data2[i] - n) <= 0.5:
                    del self.data1[i]
                    del self.data2[i]
                    del self.data3[i]
                    break
            plt.clf()

    #绘图函数的增加功能, 蓝色锥桶增加函数
    def Matlab_Add_Blue(self):
        self.Refresh_Loop()
        while (1):
            self.Matlab_Drawing()
            try:
                [(m, n)] = plt.ginput(1)
            except ValueError:
                time.sleep(0.01)
                [(m, n)] = plt.ginput(1)
            plt.scatter(m, n, color='blue', marker='.')
            self.data1.append(m)
            self.data2.append(n)
            self.data3.append(1)
            plt.clf()

    #绘图函数的增加功能, 红色锥桶增加函数
    def Matlab_Add_Red(self):
        self.Refresh_Loop()
        while (1):
            self.Matlab_Drawing()
            try:
                [(m, n)] = plt.ginput(1)
            except ValueError:
                time.sleep(0.01)
                [(m, n)] = plt.ginput(1)
            plt.scatter(m, n, color='red', marker='.')
            self.data1.append(m)
            self.data2.append(n)
            self.data3.append(2)
            plt.clf()

    # 添加自定义坐标锥筒
    def Matlab_Add(self):
        self.Matlab_Drawing()
        try:
            m = float(self.Input_x.get())
            n = float(self.Input_y.get())
        except ValueError or UnboundLocalError:
            self.root=Tk()
            self.Output_Check = Button(self.root, text='\n格式有误\n',command=self.root.destroy)
            self.Output_Check.pack()
        # 避免重复添加
        for i in range(len(self.data3)):
            if m == self.data1[i] and n == self.data2[i]:
                return
        if self.Input_type.get() == 'red':
            plt.scatter(m, n, color='red', marker='.')
            self.data1.append(m)
            self.data2.append(n)
            self.data3.append(2)
        else:
            plt.scatter(m, n, color='blue', marker='.')
            self.data1.append(m)
            self.data2.append(n)
            self.data3.append(1)
        plt.clf()
        self.Matlab_Drawing()
        plt.show()

    # 绘图函数的查看功能
    def Matlab_Check(self):
        self.Refresh_Loop()
        while (1):
            self.Matlab_Drawing()
            try:
                [(m, n)] = plt.ginput(1)
            except ValueError:
                time.sleep(0.1)
                continue
            for i in range(len(self.data3)-1):
                if abs(self.data1[i] - m) <= 0.5 and abs(self.data2[i] - n) <= 0.5:
                    if self.data3[i] == 1:
                        color = 'blue'
                    else:
                        color = 'red'
                    self.Txt.delete(1.0, 'end')
                    self.Txt.insert(
                        'end', 'x='+str(self.data1[i])+'\n'+'y='+str(self.data2[i])+'\n'+color)
                    break
            plt.clf()

    # 刷新
    def Refresh_Loop(self):
        plt.close()
        self.in_loop = 0
        self.Matlab_Drawing()
    
    def Refresh(self):
        plt.close()
        self.in_loop = 0
        self.Matlab_Drawing()
        plt.show()
        
    # 彻底退出
    def All_Close(self):
        plt.close()
        while (1):
            self.root.quit()
            self.root.destroy()

    # 导出
    def Out_Data(self):
        # 行号列号从1开始
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = u"导出数据"
        for i in range(len(self.data1)):
            ws.cell(row=i+1, column=1).value = self.data1[i]
            ws.cell(row=i+1, column=2).value = self.data2[i]
            ws.cell(row=i+1, column=3).value = self.data3[i]
        wb.save('Output.xlsx')
        #给个信
        self.root=Tk()
        self.Output_Check = Button(self.root, text='\n导出成功\n',command=self.root.destroy)
        self.Output_Check.pack()

# 主函数
def main():
    a = TrackEditor()
    a.root.mainloop()
# 执行机构
if __name__ == "__main__":
    main()

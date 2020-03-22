#中国赛地图绘制.py
#Use Tab to ident (BAD HABIT BUT ...)
#使用Tab来缩进（坏习惯。。。）
import time
import sys
from openpyxl import load_workbook
from tkinter import messagebox
import matplotlib.pyplot as plt
from tkinter import *

data1=[]
data2=[]
data3=[]

def Load_Data(cell_data_1,cell_data_2,cell_data_3):
    workbook = load_workbook('Map_Test.xlsx')    #找到需要xlsx文件的位置
    booksheet = workbook.active                 #获取当前活跃的sheet,默认是第一个sheet

    #如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workbook.get_sheet_names()  # 从名称获取sheet
    # booksheet = workbook.get_sheet_by_name(sheets[0])

    #获取sheet页的行数据
    rows = booksheet.rows
    #获取sheet页的列数据
    columns = booksheet.columns

    i = 0
    # 迭代所有的行
    for row in rows:
        i = i + 1
        line = [col.value for col in row]
        cell_data_1.append( booksheet.cell(row=i, column=1).value )              #获取第i行1列的数据
        cell_data_2.append( booksheet.cell(row=i, column=2).value )              #获取第i行2列的数据
        cell_data_3.append( booksheet.cell(row=i, column=3).value )              #获取第i行3列的数据

#绘图函数
def Matlab_Drawing():
    for i in range( len(data3) ):
        if data3[i] == 1:
            plt.scatter(data1[i],data2[i],color='blue',marker='.')
        else:
            plt.scatter(data1[i],data2[i],color='red',marker='.')

#绘图函数的删除功能
def Matlab_Delete():
    while (1): 
        Matlab_Drawing()
        try:
            [(m,n)]=plt.ginput(1)
        except ValueError:
            time.sleep(0.01)
            [(m,n)]=plt.ginput(1)
        
        for i in range( len(data3)-1 ):
            if abs(data1[i] - m) <= 0.5 and abs(data2[i] - n) <= 0.5:
                del data1[i]
                del data2[i]
                del data3[i]
                break
        plt.clf()

#绘图函数的增加功能
def Matlab_Add():
    while (1):
        Matlab_Drawing()
        try:
            [(m,n)]=plt.ginput(1)
        except ValueError:
            time.sleep(0.01)
            [(m,n)]=plt.ginput(1)

        plt.scatter(m,n,color='blue',marker='.')
        data1.append(m)
        data2.append(n)
        data3.append(1)

#Matlab的查看函数
def Matlab_Check():
    temp=0
    while (1): 
        if temp == 0:
            Matlab_Drawing()
            temp+=1
        try:
            [(m,n)]=plt.ginput(1)
        except ValueError:
            continue
        for i in range( len(data3)-1 ):
            if abs(data1[i] - m) <= 0.5 and abs(data2[i] - n) <= 0.5:
                print(data1[i],data2[i],data3[i])
                break
#主函数
def main():
    Load_Data(data1,data2,data3)

    root = Tk()
    root.title('赛道编辑器')
    label = Label(root)
    label['text'] = '赛道编辑器'
    label.pack()
    #####
    Check = Button(root)
    Check['text'] = '查看锥桶位置'
    Check['command'] = Matlab_Check
    Check.pack()
    #####
    Add = Button(root)
    Add['text'] = '  增加锥桶  '
    Add['command'] = Matlab_Add
    Add.pack()
    #####
    Delete = Button(root)
    Delete['text'] = '  删除锥桶  '
    Delete['command'] = Matlab_Delete
    Delete.pack()
    #####
    b = Button(root, text='退出', command=root.quit)
    b.pack()
    #####
    root.mainloop()

#执行机构
if __name__ == "__main__":
    main()
